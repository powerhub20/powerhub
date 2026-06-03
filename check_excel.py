from openpyxl import load_workbook

wb = load_workbook("C:\\Users\\Usuário-Pc\\Desktop\\PowerHub\\VENDAS.xlsx")
ws = wb['2026']

# Linha 14 é Junho
faturamento_junho = ws['B14'].value
anuncio_junho = ws['C14'].value

print(f"Faturamento Junho (B14): {faturamento_junho}")
print(f"Anuncio Junho (C14): {anuncio_junho}")

print("\nTodas as vendas de 2026:")
for row in range(9, 21):
    mes = ws[f'A{row}'].value
    valor = ws[f'B{row}'].value
    print(f"Linha {row} ({mes}): {valor}")
