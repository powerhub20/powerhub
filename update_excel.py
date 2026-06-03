from openpyxl import load_workbook

wb = load_workbook("C:\\Users\\Usuário-Pc\\Desktop\\PowerHub\\VENDAS.xlsx")
ws = wb['2026']

# Atualizar Junho (B14) de 1100 para 1770
ws['B14'] = 1770

# Salvar
wb.save("C:\\Users\\Usuário-Pc\\Desktop\\PowerHub\\VENDAS.xlsx")
print("[OK] Junho atualizado de 1100 para 1770!")
print("[OK] Arquivo salvo com sucesso!")

# Verificar
wb2 = load_workbook("C:\\Users\\Usuário-Pc\\Desktop\\PowerHub\\VENDAS.xlsx")
ws2 = wb2['2026']
novo_valor = ws2['B14'].value
print(f"[OK] Confirmacao: Junho agora eh {novo_valor}")
